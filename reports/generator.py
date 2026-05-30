import os
import json
import logging
from datetime import datetime
from io import BytesIO
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

logger = logging.getLogger("smart_routing.reports.generator")

class NumberedCanvas(canvas.Canvas):
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count: int):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#4b5563"))
        
        if self._pageNumber > 1:
            self.setStrokeColor(colors.HexColor("#e5e7eb"))
            self.setLineWidth(0.5)
            self.line(54, 738, 558, 738)
            self.drawString(54, 744, "MOBILIDADE INTELIGENTE - RELATÓRIO EXECUTIVO DE OTIMIZAÇÃO DE ROTAS")
            
        self.setStrokeColor(colors.HexColor("#e5e7eb"))
        self.setLineWidth(0.5)
        self.line(54, 50, 558, 50)
        
        page_str = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(558, 38, page_str)
        self.drawString(54, 38, "CONFIDENCIAL - RELATORIO DE MOBILIDADE LOGISTICA")
        self.restoreState()


class ReportGenerator:
    
    @staticmethod
    def generate_excel_report(
        locations: List[Dict[str, Any]],
        traffic: str,
        weather: str,
        metrics: List[Dict[str, Any]],
        legs_data: Dict[str, List[Dict[str, Any]]]
    ) -> BytesIO:
        logger.info("Constructing styled Excel route report...")
        wb = Workbook()
        
        font_family = "Segoe UI"
        title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
        header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        data_font = Font(name=font_family, size=10)
        bold_font = Font(name=font_family, size=10, bold=True)
        
        dark_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
        light_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        ws1 = wb.active
        ws1.title = "Resumo do Trajeto"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.merge_cells("A1:D2")
        ws1["A1"] = "METRICAS DE OTIMIZACAO DE ROTAS"
        ws1["A1"].font = title_font
        ws1["A1"].fill = dark_fill
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        
        ws1.append([])
        ws1.append([])
        
        ws1.append(["Atributo da Metrica", "Valor", "Detalhes Operacionais"])
        for col_idx in range(1, 4):
            cell = ws1.cell(row=5, column=col_idx)
            cell.font = header_font
            cell.fill = dark_fill
            cell.alignment = Alignment(horizontal="center")
            
        traffic_pt_map = {"Low": "Baixo", "Medium": "Médio", "High": "Alto", "Peak Hour": "Horário de Pico"}
        weather_pt_map = {"Sunny": "Ensolarado", "Rainy": "Chuvoso", "Snowy": "Nevando", "Stormy": "Tempestuoso"}
        traffic_pt = traffic_pt_map.get(traffic, traffic)
        weather_pt = weather_pt_map.get(weather, weather)
 
        ws1.append(["Latitude de Partida", locations[0]["lat"], "Coordenada de origem"])
        ws1.append(["Longitude de Partida", locations[0]["lng"], "Coordenada de origem"])
        ws1.append(["Quantidade de Entregas", len(locations) - 1, "Número de residências visitadas"])
        ws1.append(["Intensidade do Trânsito", traffic_pt, "Escala de trânsito aplicada"])
        ws1.append(["Clima Simulado", weather_pt, "Fator climático de ajuste de velocidade"])
        
        for r_idx in range(6, 11):
            ws1.cell(row=r_idx, column=1).font = bold_font
            ws1.cell(row=r_idx, column=2).font = data_font
            ws1.cell(row=r_idx, column=3).font = data_font
            for c_idx in range(1, 4):
                ws1.cell(row=r_idx, column=c_idx).border = thin_border
                
        ws2 = wb.create_sheet(title="Desempenho dos Veículos")
        ws2.views.sheetView[0].showGridLines = True
        
        headers2 = ["Modo de Veículo", "Distância Total (km)", "Duração da Viagem (horas)", "Combustível Consumido (L)", "Custo de Combustível (R$)", "Pegada de CO2 (g)", "Calorias Queimadas (kcal)", "Eco-Score (100)"]
        ws2.append(headers2)
        for col_idx, h in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = dark_fill
            cell.alignment = Alignment(horizontal="center")
            
        for idx, m in enumerate(metrics, 2):
            ws2.append([
                m["vehicle"], m["distance_km"], m["time_hours"],
                m["fuel_liters"], m["fuel_cost_usd"], m["co2_emissions_g"],
                m["calories_burned"], m["sustainability_score"]
            ])
            ws2.cell(row=idx, column=2).number_format = "0.00"
            ws2.cell(row=idx, column=3).number_format = "0.00"
            ws2.cell(row=idx, column=4).number_format = "0.00"
            ws2.cell(row=idx, column=5).number_format = "$#,##0.00"
            ws2.cell(row=idx, column=6).number_format = "#,##0.0"
            ws2.cell(row=idx, column=7).number_format = "0.0"
            
            for c_idx in range(1, 9):
                cell = ws2.cell(row=idx, column=c_idx)
                cell.font = data_font
                cell.border = thin_border
                if idx % 2 == 0:
                    cell.fill = light_fill

        ws3 = wb.create_sheet(title="Trechos da Rota")
        ws3.views.sheetView[0].showGridLines = True
        
        headers3 = ["Modo de Veículo", "Origem do Trecho", "Destino do Trecho", "Distância (km)"]
        ws3.append(headers3)
        for col_idx, h in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = dark_fill
            cell.alignment = Alignment(horizontal="center")
            
        r_counter = 2
        for vehicle, legs in legs_data.items():
            for leg in legs:
                ws3.append([
                    vehicle, leg["from_node"], leg["to_node"], leg["distance_km"]
                ])
                ws3.cell(row=r_counter, column=4).number_format = "0.00"
                for c in range(1, 5):
                    cell = ws3.cell(row=r_counter, column=c)
                    cell.font = data_font
                    cell.border = thin_border
                r_counter += 1
                
        for ws in [ws1, ws2, ws3]:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def generate_pdf_report(
        locations: List[Dict[str, Any]],
        traffic: str,
        weather: str,
        metrics: List[Dict[str, Any]],
        optimal_order: List[int],
        recommendation: Dict[str, Any]
    ) -> BytesIO:
        logger.info("Generating styled PDF routing report with ReportLab...")
        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=letter,
            rightMargin=54, leftMargin=54, topMargin=54, bottomMargin=54
        )
        
        slate_dark = colors.HexColor("#111827")
        emerald_main = colors.HexColor("#10B981")
        grey_light = colors.HexColor("#f3f4f6")
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CoverTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=24,
            textColor=slate_dark,
            spaceAfter=15,
            leading=28
        )
        
        subtitle_style = ParagraphStyle(
            'CoverSubtitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=12,
            textColor=colors.HexColor("#4b5563"),
            spaceAfter=30,
            leading=16
        )
        
        heading1_style = ParagraphStyle(
            'ReportHeading1',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=slate_dark,
            spaceBefore=15,
            spaceAfter=10,
            keepWithNext=True
        )
        
        body_style = ParagraphStyle(
            'ReportBody',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor("#1f2937"),
            spaceAfter=8,
            leading=14
        )
        
        recommendation_style = ParagraphStyle(
            'RecBlock',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.HexColor("#065f46"),
            backColor=colors.HexColor("#ecfdf5"),
            borderColor=colors.HexColor("#a7f3d0"),
            borderWidth=1,
            borderPadding=10,
            spaceAfter=15,
            leading=14
        )

        story = []
        
        traffic_pt_map = {"Low": "Baixo", "Medium": "Médio", "High": "Alto", "Peak Hour": "Horário de Pico"}
        weather_pt_map = {"Sunny": "Ensolarado", "Rainy": "Chuvoso", "Snowy": "Nevando", "Stormy": "Tempestuoso"}
        traffic_pt = traffic_pt_map.get(traffic, traffic)
        weather_pt = weather_pt_map.get(weather, weather)

        story.append(Spacer(1, 100))
        story.append(Paragraph("OTIMIZAÇÃO DE ROTAS INTELIGENTE", title_style))
        story.append(Paragraph("Relatório de Mobilidade Tática e Eficiência Logística", subtitle_style))
        story.append(Spacer(1, 20))
        
        story.append(Table(
            [[""]],
            colWidths=[504],
            rowHeights=[6],
            style=TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), emerald_main),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ])
        ))
        story.append(Spacer(1, 200))
        
        story.append(Paragraph(f"<b>Gerado em:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", body_style))
        story.append(Paragraph("<b>Sistema Logístico:</b> Motor de Mobilidade Inteligente v1.0", body_style))
        story.append(Paragraph("<b>Escopo:</b> 1 Origem + 3 Residências de Entrega", body_style))
        story.append(PageBreak())
        
        story.append(Paragraph("1. Resumo Executivo e Configuração", heading1_style))
        story.append(Paragraph(
            "Este relatório documenta a avaliação logística de múltiplos destinos. "
            "O sistema calcula a sequência ideal de visitas (resolução do Problema do Caixeiro Viajante) "
            "para minimizar a distância e o tempo geral, comparando modos de transporte motorizados e não motorizados "
            "sob parâmetros de trânsito e condições climáticas.",
            body_style
        ))
        
        meta_data = [
            ["Configurações de Roteamento", "Valor", "Notas"],
            ["Coordenadas de Origem", f"{locations[0]['lat']:.5f}, {locations[0]['lng']:.5f}", "Ponto base inicial"],
            ["Condições do Tempo", weather_pt, "Ajusta a segurança e velocidade de viagem"],
            ["Congestionamento de Trânsito", traffic_pt, "Fator de redução de velocidade aplicado"],
            ["Sequência de Visitas", " -> ".join([f"Entrega {i+1}" for i in optimal_order]), "Ordem de visita otimizada"]
        ]
        t_meta = Table(meta_data, colWidths=[150, 150, 204])
        t_meta.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), slate_dark),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e5e7eb")),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, grey_light])
        ]))
        story.append(t_meta)
        story.append(Spacer(1, 20))
        
        story.append(Paragraph("2. Comparativo de Métricas por Veículo", heading1_style))
        story.append(Paragraph(
            "A tabela abaixo detalha valores de distância, duração, consumo de combustível, custo, pegada de carbono e calorias. "
            "Observe como os atrasos de trânsito afetam o tempo de viagem dos carros, enquanto as bicicletas permanecem altamente previsíveis, porém mais lentas.",
            body_style
        ))
        
        metrics_table_data = [["Modo de Veículo", "Dist. (km)", "Tempo (min)", "Combustível (L)", "Custo (R$)", "CO2 (g)", "Calorias (kcal)"]]
        for m in metrics:
            metrics_table_data.append([
                m["vehicle"],
                f"{m['distance_km']:.2f}",
                f"{m['time_hours'] * 60.0:.1f}",
                f"{m['fuel_liters']:.2f}",
                f"R$ {m['fuel_cost_usd']:.2f}",
                f"{m['co2_emissions_g']:.0f}",
                f"{m['calories_burned']:.0f}"
            ])
            
        t_metrics = Table(metrics_table_data, colWidths=[100, 70, 74, 60, 60, 70, 70])
        t_metrics.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), slate_dark),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e5e7eb")),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, grey_light])
        ]))
        story.append(t_metrics)
        story.append(Spacer(1, 25))
        
        story.append(Paragraph("3. Recomendação do Sistema IA", heading1_style))
        
        traffic_pt_map = {"Low": "Baixo", "Medium": "Médio", "High": "Alto", "Peak Hour": "Horário de Pico"}
        weather_pt_map = {"Sunny": "Ensolarado", "Rainy": "Chuvoso", "Snowy": "Nevando", "Stormy": "Tempestuoso"}
        traffic_pt = traffic_pt_map.get(traffic, traffic)
        weather_pt = weather_pt_map.get(weather, weather)
        
        rec_txt = (
            f"<b>🏆 MELHOR OPÇÃO: {recommendation['best_vehicle'].upper()}</b><br/><br/>"
            f"<b>Racional Operacional:</b> {recommendation['reason']}<br/><br/>"
            f"<b>Avaliação de Tempo vs Custo:</b> A ordem ideal da rota satisfaz as restrições do Problema do Caixeiro Viajante. "
            f"Sob o trânsito atual ({traffic_pt}) e clima ({weather_pt}), utilizar o(a) {recommendation['best_vehicle']} otimiza o equilíbrio "
            f"entre urgência de cronograma, restrições financeiras e metas de sustentabilidade."
        )
        story.append(Paragraph(rec_txt, recommendation_style))
        
        doc.build(story, canvasmaker=NumberedCanvas)
        buf.seek(0)
        return buf
#A
